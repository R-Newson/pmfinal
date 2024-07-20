from rest_framework import viewsets
from .models import Task
from .serializers import TaskSerializer
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .forms import CreateUserForm, LoginForm, ParetoForm
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.http import HttpResponse
import openpyxl
from openpyxl.chart import BarChart, Reference
from io import BytesIO

class TaskViewSet(viewsets.ModelViewSet):
    queryset = Task.objects.all()
    serializer_class = TaskSerializer

def homepage(request):
    return render(request, 'index.html')

def signup(request):
    form = CreateUserForm()
    if request.method == "POST":
        form = CreateUserForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("login")
    context = {'registerform': form}
    return render(request, 'signup.html', context=context)

def user_login(request):
    form = LoginForm()
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                auth_login(request, user)
                return redirect('index')
    context = {'loginform': form}
    return render(request, 'login.html', context=context)

def pareto_chart_view(request):
    if request.method == 'POST':
        form = ParetoForm(request.POST)
        if form.is_valid():
            complaint_types = [ct.strip() for ct in form.cleaned_data['complaint_type'].split(',')]
            number_of_complaints = list(map(int, form.cleaned_data['number_of_complaints'].split(',')))

            # Create a workbook and select the active worksheet
            wb = openpyxl.Workbook()
            ws = wb.active

            # Write data to the worksheet
            ws.append(['Complaint Type', 'Number of Complaints'])
            for ct, nc in zip(complaint_types, number_of_complaints):
                ws.append([ct, nc])

            # Sort data by 'Number of Complaints'
            sorted_data = sorted(zip(complaint_types, number_of_complaints), key=lambda x: x[1], reverse=True)
            for idx, (ct, nc) in enumerate(sorted_data, start=2):
                ws[f'A{idx}'] = ct
                ws[f'B{idx}'] = nc

            # Calculate cumulative percentage
            total_complaints = sum(number_of_complaints)
            cumulative_percentage = []
            cumulative_sum = 0
            for nc in [row[1] for row in sorted_data]:
                cumulative_sum += nc
                cumulative_percentage.append(cumulative_sum / total_complaints * 100)

            for idx, cp in enumerate(cumulative_percentage, start=2):
                ws[f'C{idx}'] = cp

            # Create a bar chart
            bar_chart = BarChart()
            data = Reference(ws, min_col=2, min_row=1, max_row=len(complaint_types) + 1)
            categories = Reference(ws, min_col=1, min_row=2, max_row=len(complaint_types) + 1)
            bar_chart.add_data(data, titles_from_data=True)
            bar_chart.set_categories(categories)
            bar_chart.y_axis.title = 'Number of Complaints'
            bar_chart.x_axis.title = 'Complaint Type'

            # Color bars above 80% differently
            for i, nc in enumerate([row[1] for row in sorted_data]):
                if nc > 80:
                    # Apply red fill to bars above 80%
                    bar_chart.series[0].graphicalProperties.solidFill = 'FF0000'

            ws.add_chart(bar_chart, 'E5')

            # Save the workbook to a BytesIO object
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            # Serve the Excel file as a response
            response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=pareto.xlsx'
            return response
    else:
        form = ParetoForm()

    return render(request, 'pareto_chart.html', {'form': form})

@login_required
def user_logout(request):
    if request.method == 'POST':
        auth_logout(request)
        return redirect('index')

@login_required
def index(request):
    return render(request, 'index.html')

@login_required
def about(request):
    return render(request, 'about.html')

@login_required
def projects(request):
    return render(request, 'projects.html')

@login_required
def quality_centre(request):
    return render(request, 'quality_centre.html')

@login_required
def analytics(request):
    return render(request, 'analytics.html')

@login_required
def pareto(request):
    return render(request, 'pareto_chart.html')
